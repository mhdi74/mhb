import io
import re
import math
import base64
import os
import hashlib
import secrets
from copy import copy
from datetime import datetime, timedelta
from typing import List, Optional

import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
from pyproj import Transformer
import requests

from fastapi import FastAPI, HTTPException, Depends, status
from fastapi.security import OAuth2PasswordBearer
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

# کتابخانه‌های دیتابیس
from sqlalchemy import create_engine, Column, Integer, String, Float, Text, Boolean, DateTime, ForeignKey
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
from jose import JWTError, jwt

# --- تنظیمات متغیرهای محیطی ---
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://postgres:password@localhost:5432/postgres")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

SECRET_KEY = os.getenv("SECRET_KEY", "cadastre_super_secret_key_123456789")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # اعتبار توکن: ۷ روز

BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
ADMIN_CHAT_ID = os.getenv("ADMIN_CHAT_ID", "")
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")  # نام کاربری مدیر سامانه

# --- تنظیمات SQLAlchemy و دیتابیس ---
engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# --- مدل‌های دیتابیس (ORM) ---
class UserDB(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String(50), unique=True, index=True, nullable=False)
    password_hash = Column(String(255), nullable=False)
    full_name = Column(String(100), nullable=True)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)

class CertificateDB(Base):
    __tablename__ = "certificates"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"))
    applicant_name = Column(String(100))
    national_id = Column(String(20))
    total_area = Column(Float)
    zone = Column(Integer)
    coords_text = Column(Text)
    created_at = Column(DateTime, default=datetime.utcnow)

# ساخت اتوماتیک جداول در دیتابیس
Base.metadata.create_all(bind=engine)

# --- تنظیمات امنیت و هشینگ استاندارد پایتون (بدون وابسته خارجی) ---
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="login")

def get_password_hash(password: str) -> str:
    salt = secrets.token_hex(16)
    pwd_hash = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 100000).hex()
    return f"{salt}${pwd_hash}"

def verify_password(plain_password: str, hashed_password: str) -> bool:
    try:
        salt, stored_hash = hashed_password.split('$')
        pwd_hash = hashlib.pbkdf2_hmac('sha256', plain_password.encode('utf-8'), salt.encode('utf-8'), 100000).hex()
        return secrets.compare_digest(pwd_hash, stored_hash)
    except Exception:
        return False

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

# تابع گرفتن Session دیتابیس
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# تابع بررسی توکن و گرفتن کاربر جاری
async def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="توکن نامعتبر یا منقضی شده است.",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
        
    user = db.query(UserDB).filter(UserDB.username == username).first()
    if user is None or not user.is_active:
        raise credentials_exception
    return user

# تابع ارسال پیام به تلگرام مدیر
def send_admin_notification(message_text: str):
    if not BOT_TOKEN or not ADMIN_CHAT_ID:
        return
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        payload = {
            "chat_id": ADMIN_CHAT_ID,
            "text": message_text,
            "parse_mode": "HTML"
        }
        requests.post(url, json=payload, timeout=5)
    except Exception as e:
        print(f"خطا در ارسال پیام به مدیر: {e}")

# --- مدل‌های Pydantic ---
class RegisterRequest(BaseModel):
    username: str
    password: str
    full_name: Optional[str] = None

class LoginRequest(BaseModel):
    username: str
    password: str

class CadastreRequest(BaseModel):
    is_telegram: bool = False
    user_id: Optional[int] = None
    applicant_name: str
    national_id: str
    building_area: float = 0.0
    zone: int = 39
    coords_text: str
    ranges: List[dict]

# --- ساخت برنامه FastAPI ---
app = FastAPI(title="Cadastre API Engine with Auth")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- توابع محاسباتی کاداستر ---
def parse_smart_coordinates(text_input: str) -> pd.DataFrame:
    parsed_points = []
    lines = text_input.strip().split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        if "X=" in line or "Y=" in line:
            try:
                x_match = re.search(r'X\s*=\s*([0-9\.]+)', line)
                y_match = re.search(r'Y\s*=\s*([0-9\.]+)', line)
                if x_match and y_match:
                    parsed_points.append((float(x_match.group(1)), float(y_match.group(1))))
                    continue
            except Exception:
                pass

        tokens = re.split(r'[\s,;\t\-]+', line)
        numeric_tokens = []
        for t in tokens:
            t_clean = re.sub(r'[^\d\.]', '', t)
            try:
                if t_clean:
                    numeric_tokens.append(float(t_clean))
            except ValueError:
                continue
        
        if len(numeric_tokens) == 2:
            parsed_points.append((numeric_tokens[0], numeric_tokens[1]))
        elif len(numeric_tokens) >= 3:
            if numeric_tokens[0] < 10000: 
                parsed_points.append((numeric_tokens[1], numeric_tokens[2]))
            else:
                parsed_points.append((numeric_tokens[0], numeric_tokens[1]))
            
    if not parsed_points:
        raise ValueError("هیچ مختصات معتبری در متن ورودی پیدا نشد!")
        
    return pd.DataFrame(parsed_points, columns=['X', 'Y'])

def reorder_polygon_clockwise_from_north(df: pd.DataFrame) -> pd.DataFrame:
    x = df['X'].to_numpy()
    y = df['Y'].to_numpy()
    
    edges_sum = np.sum((x[1:] - x[:-1]) * (y[1:] + y[:-1]))
    edges_sum += (x[0] - x[-1]) * (y[0] + y[-1])
    
    if edges_sum < 0: 
        x = np.flip(x)
        y = np.flip(y)
        
    north_idx = np.argmax(y)
    x_ordered = np.roll(x, -north_idx)
    y_ordered = np.roll(y, -north_idx)
    
    return pd.DataFrame({'X': x_ordered, 'Y': y_ordered})

def calculate_area_and_centroid(df: pd.DataFrame, zone: int = 39):
    x = df['X'].values
    y = df['Y'].values
    
    if x[0] != x[-1] or y[0] != y[-1]:
        x = np.append(x, x[0])
        y = np.append(y, y[0])
    
    area_val = 0.5 * np.sum(x[:-1] * y[1:] - x[1:] * y[:-1])
    abs_area = abs(area_val)
    
    cx = (1 / (6 * area_val)) * np.sum((x[:-1] + x[1:]) * (x[:-1] * y[1:] - x[1:] * y[:-1]))
    cy = (1 / (6 * area_val)) * np.sum((y[:-1] + y[1:]) * (x[:-1] * y[1:] - x[1:] * y[:-1]))
    
    transformer = Transformer.from_crs(f"epsg:326{zone}", "epsg:4326")
    lat, lon = transformer.transform(cx, cy)
    
    return float(cx), float(cy), float(lat), float(lon), float(abs_area)

def copy_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.alignment = copy(source_cell.alignment)

def safe_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                break
    else:
        cell.value = value

def get_neighbor_text(p1: int, ranges_list: list) -> str:
    for row in ranges_list:
        s = row.get('راس ابتدا', 1)
        e = row.get('راس انتها', 2)
        mtype = row.get('نوع مرز', 'مرز')
        name = row.get('نام مجاور', '---')
        if s <= p1 < e or (e < s and (p1 >= s or p1 < e)):
            return f"مرزی است \\ {name}" if mtype == "مرز" else f"{mtype}ی است \\ {name}"
    return "---"

def send_file_to_telegram(chat_id: int, file_bytes: bytes, filename: str):
    if not BOT_TOKEN:
        return
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
    files = {'document': (filename, file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    data = {'chat_id': chat_id}
    requests.post(url, data=data, files=files)

# --- مسیرهای API (Endpoints) ---

@app.get("/")
@app.head("/")
async def root():
    return {"status": "online", "message": "سامانه کاداستر همراه با دیتابیس و احراز هویت فعال است."}

@app.post("/register")
async def register(req: RegisterRequest, db: Session = Depends(get_db)):
    try:
        existing_user = db.query(UserDB).filter(UserDB.username == req.username).first()
        if existing_user:
            raise HTTPException(status_code=400, detail="این نام کاربری قبلاً ثبت شده است.")
        
        hashed_pwd = get_password_hash(req.password)
        new_user = UserDB(username=req.username, password_hash=hashed_pwd, full_name=req.full_name)
        db.add(new_user)
        db.commit()
        db.refresh(new_user)

        # ارسال اعلان ثبت نام کاربر جدید به مدیر در تلگرام
        admin_msg = (
            f"👤 <b>ثبت‌نام کاربر جدید در سامانه!</b>\n\n"
            f"🔹 نام کاربری: <code>{new_user.username}</code>\n"
            f"🔹 نام و خانوادگی: {new_user.full_name or 'ثبت نشده'}\n"
            f"📅 تاریخ: {new_user.created_at.strftime('%Y-%m-%d %H:%M:%S')}"
        )
        send_admin_notification(admin_msg)

        return {"success": True, "message": "ثبت نام با موفقیت انجام شد."}
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"خطای داخلی سرور: {str(e)}")

@app.post("/login")
async def login(req: LoginRequest, db: Session = Depends(get_db)):
    try:
        user = db.query(UserDB).filter(UserDB.username == req.username).first()
        if not user or not verify_password(req.password, user.password_hash):
            raise HTTPException(status_code=400, detail="نام کاربری یا رمز عبور اشتباه است.")
        
        access_token = create_access_token(data={"sub": user.username})
        # مشخص کردن دسترسی مدیر بودن
        is_admin = (user.username.lower() == ADMIN_USERNAME.lower())
        
        return {
            "access_token": access_token, 
            "token_type": "bearer",
            "is_admin": is_admin
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"خطای داخلی سرور: {str(e)}")

@app.post("/process")
async def process_cadastre(
    req: CadastreRequest, 
    current_user: UserDB = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    try:
        # ۱. محاسبات هندسی
        df_raw = parse_smart_coordinates(req.coords_text)
        df = reorder_polygon_clockwise_from_north(df_raw)
        num_points = len(df)
        cx, cy, lat, lon, total_area = calculate_area_and_centroid(df, req.zone)

        # تبدیل قطعی متغیر numpy به float خالص پایتون
        total_area_val = float(np.asarray(total_area).item()) if hasattr(total_area, 'item') else float(total_area)

        # ۲. پردازش اکسل ۱
        wb1 = load_workbook("temp1.xlsx")
        ws1 = wb1.active
        start_row1, template_capacity1 = 7, 16
        
        if num_points > template_capacity1:
            diff1 = num_points - template_capacity1
            ws1.insert_rows(start_row1 + template_capacity1, amount=diff1)
            for r in range(start_row1 + template_capacity1, start_row1 + num_points):
                for c in range(1, 10):
                    copy_style(ws1.cell(row=start_row1 + template_capacity1 - 1, column=c), ws1.cell(row=r, column=c))

        for i in range(num_points):
            p1_id = i + 1
            dist = math.sqrt((df.iloc[(i+1)%num_points].X - df.iloc[i].X)**2 + (df.iloc[(i+1)%num_points].Y - df.iloc[i].Y)**2)
            adj_text = get_neighbor_text(p1_id, req.ranges)
            curr_row = start_row1 + i
            safe_write(ws1, curr_row, 1, p1_id)
            safe_write(ws1, curr_row, 2, round(dist, 4))
            safe_write(ws1, curr_row, 4, adj_text)

        safe_write(ws1, 3, 4, f"{lon:.7f}°") 
        safe_write(ws1, 4, 4, f"{lat:.7f}°") 
        
        output1 = io.BytesIO()
        wb1.save(output1)
        bytes1 = output1.getvalue()

        # ۳. پردازش اکسل ۲
        wb2 = load_workbook("temp2.xlsx")
        ws2 = wb2.active
        
        safe_write(ws2, 2, 3, req.applicant_name) 
        safe_write(ws2, 3, 3, req.national_id)    
        safe_write(ws2, 5, 3, round(total_area_val, 2)) 
        safe_write(ws2, 6, 3, req.building_area)  
        safe_write(ws2, 8, 3, req.zone)           
        safe_write(ws2, 9, 3, num_points)     

        start_row2, template_capacity2 = 11, 14
        if num_points > template_capacity2:
            diff2 = num_points - template_capacity2
            ws2.insert_rows(start_row2 + template_capacity2, amount=diff2)
            for r in range(start_row2 + template_capacity2, start_row2 + num_points):
                for c in range(1, 6):
                    copy_style(ws2.cell(row=start_row2 + template_capacity2 - 1, column=c), ws2.cell(row=r, column=c))

        for i in range(num_points):
            p1_id = i + 1
            dist = math.sqrt((df.iloc[(i+1)%num_points].X - df.iloc[i].X)**2 + (df.iloc[(i+1)%num_points].Y - df.iloc[i].Y)**2)
            adj_text = get_neighbor_text(p1_id, req.ranges)
            curr_row = start_row2 + i
            safe_write(ws2, curr_row, 1, p1_id)                  
            safe_write(ws2, curr_row, 2, df.iloc[i].X)           
            safe_write(ws2, curr_row, 3, df.iloc[i].Y)           
            safe_write(ws2, curr_row, 4, adj_text)               
            safe_write(ws2, curr_row, 5, round(dist, 4))         

        output2 = io.BytesIO()
        wb2.save(output2)
        bytes2 = output2.getvalue()

        # ۴. ذخیره لاگ در دیتابیس به صورت کاملاً ایزوله (حتی اگر دیتابیس ارور دهد، خروجی اکسل کاربر متوقف نمی‌شود)
        try:
            cert_record = CertificateDB(
                user_id=current_user.id,
                applicant_name=req.applicant_name,
                national_id=req.national_id,
                total_area=total_area_val,
                zone=int(req.zone),
                coords_text=req.coords_text
            )
            db.add(cert_record)
            db.commit()
        except Exception as db_err:
            print(f"⚠️ اشکال غیربحرانی دیتابیس (خروجی متوقف نشد): {db_err}")
            db.rollback()

        # ۵. ارسال اعلان صدور نقشه به مدیر در تلگرام
        admin_msg = (
            f"📜 <b>صدور گواهی کاداستر جدید!</b>\n\n"
            f"👤 صادرکننده: <code>{current_user.username}</code> ({current_user.full_name or '---'})\n"
            f"👨‍💼 متقاضی: <b>{req.applicant_name}</b>\n"
            f"🪪 کد ملی: <code>{req.national_id}</code>\n"
            f"📐 مساحت کل: <b>{total_area_val:,.2f} متر مربع</b>\n"
            f"🌍 زون UTM: {req.zone}\n"
            f"📍 تعداد رئوس: {num_points}"
        )
        send_admin_notification(admin_msg)

        # ۶. تحویل فایل‌ها به کاربر
        filename1 = f"Gvahi_Azla_{req.applicant_name}.xlsx"
        filename2 = f"ParcelMap_{req.applicant_name}.xlsx"

        if req.is_telegram and req.user_id:
            send_file_to_telegram(req.user_id, bytes1, filename1)
            send_file_to_telegram(req.user_id, bytes2, filename2)
            return {"success": True, "message": "فایل‌ها به تلگرام ارسال شدند."}
        else:
            return {
                "success": True,
                "files": [
                    {"filename": filename1, "content_base64": base64.b64encode(bytes1).decode('utf-8')},
                    {"filename": filename2, "content_base64": base64.b64encode(bytes2).decode('utf-8')}
                ]
            }

    except Exception as e:
        return {"success": False, "error": str(e)}

# --- مسیر دریافت گزارش‌های مدیریتی (مخصوص وب) ---
@app.get("/admin/reports")
async def get_admin_reports(
    current_user: UserDB = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # 🔒 بررسی دسترسی: فقط اگر نام کاربری با ADMIN_USERNAME برابر باشد اجازه دارد
    if current_user.username.lower() != ADMIN_USERNAME.lower():
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="دستور غیرمجاز! فقط مدیر سامانه به این بخش دسترسی دارد."
        )

    try:
        # ۱. دریافت لیست تمام کاربران
        users = db.query(UserDB).order_by(UserDB.created_at.desc()).all()
        users_list = [
            {
                "id": u.id,
                "username": u.username,
                "full_name": u.full_name or "---",
                "created_at": u.created_at.strftime("%Y-%m-%d %H:%M:%S")
            }
            for u in users
        ]

        # ۲. دریافت لیست تمام گواهی‌های صادر شده همراه مشخصات صادرکننده
        certs = db.query(
            CertificateDB.id,
            CertificateDB.applicant_name,
            CertificateDB.national_id,
            CertificateDB.total_area,
            CertificateDB.zone,
            CertificateDB.created_at,
            UserDB.username,
            UserDB.full_name
        ).join(UserDB, CertificateDB.user_id == UserDB.id).order_by(CertificateDB.created_at.desc()).all()

        certs_list = [
            {
                "id": c.id,
                "applicant_name": c.applicant_name,
                "national_id": c.national_id,
                "total_area": round(float(c.total_area), 2) if c.total_area else 0.0,
                "zone": c.zone,
                "created_at": c.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "issued_by_username": c.username,
                "issued_by_fullname": c.full_name or "---"
            }
            for c in certs
        ]

        # ۳. محاسبه مجموع مساحت‌های پردازش‌شده
        total_area_sum = sum(c["total_area"] for c in certs_list) if certs_list else 0.0

        return {
            "success": True,
            "summary": {
                "total_users": len(users_list),
                "total_certificates": len(certs_list),
                "total_area_sqm": round(total_area_sum, 2)
            },
            "users": users_list,
            "certificates": certs_list
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"خطا در دریافت گزارش‌ها: {str(e)}")
